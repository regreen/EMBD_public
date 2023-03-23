'''
Created on Feb 21, 2017

@author: Admin
'''

import openpyxl, glob, os
from varLookup import FOLDER_IN, printLog
from myFunc import sort_nicely

if __name__ == '__main__':
    folder = FOLDER_IN+'EMBD_Scoring/indiv_scored_embs/'
    folderOut = FOLDER_IN+'MachineLearningTrainingFiles/'
    fileOutAllMS = folderOut + 'all_phenotypes_MS.csv'
    fileOutAllGLS = folderOut + 'all_phenotypes_GLS.csv'
    fileNames = glob.glob('{0}*.xlsx'.format(folder))
    sort_nicely(fileNames)
    allEmbs = []
    for fileName in fileNames[:]:
        print(fileName)
#         date = fileName[-11:-5]
#         newDate = '20{y}{m}{d}'.format(m=date[:2],d=date[2:4],y=date[4:])
        wb = openpyxl.load_workbook(fileName)
        sheet = wb.get_sheet_by_name('Sheet1')
        nRows = sheet.max_row
        nCol = sheet.max_column
        nextFlag = False
        rna = None
        strain = None
        emb = None
        for r in range(4,nRows):
            if sheet.cell(row=r, column=4).value=='SUM TOTAL': nextFlag=True
            else:
                if nextFlag:
                    if isinstance(sheet.cell(row=r, column=3).value, long):
                        rna = 'EMBD{r:04}'.format(r=sheet.cell(row=r, column=3).value)
                        if sheet.cell(row=r, column=1).value[0]=='G': strain = 'GLS'
                        else: strain='MS'
                    else: rna=None
                    nextFlag = False
                if rna is not None:
                    e = sheet.cell(row=r, column=6).value
                    emb = 'Emb{e}'.format(e=e)
                    if str(e)[0]!='x':
                        tfolder = FOLDER_IN+'cropped/{r}/{s}/{e}/'.format(r=rna, s=strain, e=emb)
                        if os.path.exists(tfolder):
                            files = glob.glob('{0}*.tif'.format(tfolder))
                            if len(files)>0:
                                date = glob.glob('{0}*.tif'.format(tfolder))[0].split('_')[2]
                                desc = []
                                for j in range(11,31):
                                    v = sheet.cell(row=r, column=j).value
                                    if v is None: v=0
                                    desc.append(v)
                                allEmbs.append([strain, date, rna, emb]+desc)
                            else: printLog('{r} {l} folder is empty'.format(r=rna, l=emb))
                        else: printLog('{r} {l} does not exist'.format(r=rna, l=emb))
                    
#     csvfileMS = open(fileOutAllMS, 'wb')
#     spamwriterMS = csv.writer(csvfileMS, delimiter=',',\
#                         quotechar='|', quoting=csv.QUOTE_MINIMAL)
#     csvfileGLS = open(fileOutAllGLS, 'wb')
#     spamwriterGLS = csv.writer(csvfileGLS, delimiter=',',\
#                         quotechar='|', quoting=csv.QUOTE_MINIMAL)
#     spamwriterMS.writerow(['strain', 'date', 'Target', 'Emb']+range(1,21))
#     spamwriterGLS.writerow(['strain', 'date', 'Target', 'Emb']+range(1,21))
#     for emb in allEmbs:
#         if emb[0]=='MS': spamwriterMS.writerow(emb)
#         else: spamwriterGLS.writerow(emb)
#     csvfileMS.close()
#     csvfileGLS.close()
#     for d in range(1,21):
#         fileNameMS = folderOut + '{d:02}_MS_training.csv'.format(d=d)
#         fileNameGLS = folderOut + '{d:02}_GLS_training.csv'.format(d=d)
#         csvfileMS = open(fileNameMS, 'wb')
#         csvfileGLS = open(fileNameGLS, 'wb')
#         spamwriterMS = csv.writer(csvfileMS, delimiter=',',\
#                             quotechar='|', quoting=csv.QUOTE_MINIMAL)
#         spamwriterGLS = csv.writer(csvfileGLS, delimiter=',',\
#                             quotechar='|', quoting=csv.QUOTE_MINIMAL)
#         spamwriterMS.writerow(['strain', 'date', 'Target', 'Emb'])
#         spamwriterGLS.writerow(['strain', 'date', 'Target', 'Emb'])
#         for emb in allEmbs:
#             if emb[0]=='MS' and emb[3+d]==1: spamwriterMS.writerow(emb)
#             elif emb[0]=='GLS' and emb[3+d]==1: spamwriterGLS.writerow(emb)
#         csvfileMS.close()
#         csvfileGLS.close()